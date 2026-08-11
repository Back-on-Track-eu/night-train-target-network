"""
Shared workbook access for the ontd loaders.

Both ONTD imports read Google Sheets exported as .xlsx — the refreshed
tables from the live ONTD sheet (db/ontd/loader.py) and the curated
composition catalog from the target-network workbook
(db/ontd/composition_loader.py). The fetch, the header-name row extraction
and the cell coercion rules are identical for both, so they live here.

Column extraction is header-name based (case-insensitive): callers ask
for the DB column names they know about, and anything else on the sheet
is ignored. That is what makes the known sheet artifacts a non-issue
without special-casing — trailing Chatbase/AutoCrat columns, the leading
train_stop_id on trip_stop, diverging columns on the *_inactive tabs.
"""

import datetime
import io
import os
import urllib.request
import zipfile
from typing import Any, Optional

import openpyxl

# Formula error values, which openpyxl surfaces as plain strings. Both
# source workbooks carry these routinely and they reappear after every
# refresh, so they are normalised to NULL on read rather than fixed at
# source. Without this they would land verbatim in TEXT columns; numeric
# columns only escape by accident, via a failed parse.
EXCEL_ERRORS = frozenset(
    {"#REF!", "#VALUE!", "#N/A", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!", "#ERROR!"}
)

_BOOL_TRUE = {"TRUE", "YES", "Y", "T", "1"}
_BOOL_FALSE = {"FALSE", "NO", "N", "F", "0", ""}


def sheets_export_url(sheet_id: str) -> str:
    """Whole-workbook export of a NATIVE Google Sheet — no API key or
    credentials, the sheet only has to be link-shared
    (adapters/proposal/README.md decision 26). Native sheet ids are 44 chars.

    This endpoint converts a live sheet to .xlsx on the fly; it does NOT
    serve .xlsx files merely stored in Drive — see drive_download_url.
    """
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def drive_download_url(file_id: str) -> str:
    """Download a binary file stored in Drive — an .xlsx uploaded without
    conversion (Drive file ids are ~33 chars; such links carry
    `rtpof=true`, "retain original file type"). `confirm=t` pre-answers
    the virus-scan interstitial Drive shows for files it can't scan,
    which otherwise returns an HTML page instead of the bytes.
    """
    return (
        f"https://drive.usercontent.google.com/download"
        f"?id={file_id}&export=download&confirm=t"
    )


DRIVE_FILE = "drive_file"
NATIVE_SHEET = "native_sheet"


def workbook_url(id_var: str, kind_var: str, description: str) -> str:
    """Resolve a workbook id from the environment into a download URL.

    The ids live in backend/docker/.env (documented in .env.example
    alongside GRAPH_CACHE_PATH and EEZ_LAND_UNION_FILE_ID) rather
    than in code, so a link is rotated in one place. `kind_var` selects
    the endpoint, because the two are not interchangeable: a native sheet
    is exported on the fly and always current, while a Drive-hosted .xlsx
    is served as stored bytes with formulas frozen at upload.

    .env is loaded here rather than assuming the caller connected to the
    database first — resolution should not depend on statement order.
    """
    # Local import: connection.py pulls in psycopg2, which a workbook
    # helper has no other reason to require.
    from connection import load_env_files

    load_env_files()
    file_id = os.environ.get(id_var)
    if not file_id:
        raise SystemExit(
            f"{id_var} is not set — {description}.\n"
            f"Add it to backend/docker/.env (see .env.example), or pass a "
            f"local workbook with --xlsx."
        )
    kind = os.environ.get(kind_var, DRIVE_FILE)
    if kind == NATIVE_SHEET:
        return sheets_export_url(file_id)
    if kind != DRIVE_FILE:
        raise SystemExit(
            f"{kind_var}='{kind}' is not valid — expected "
            f"'{DRIVE_FILE}' or '{NATIVE_SHEET}'."
        )
    return drive_download_url(file_id)


def load_workbook(url: Optional[str] = None, path: Optional[str] = None):
    """Fetch from `url`, or open local `path`. A non-.xlsx response (e.g.
    an HTML "request access" page when the sheet isn't link-shared) fails
    loudly here rather than producing empty tables downstream."""
    if path:
        print(f"Loading workbook from local file: {path}")
        source: Any = path
    elif url:
        print(f"Fetching workbook from {url} ...")
        with urllib.request.urlopen(url) as response:
            source = io.BytesIO(response.read())
    else:
        raise ValueError("load_workbook needs either url or path")

    try:
        return openpyxl.load_workbook(source, data_only=True, read_only=True)
    except zipfile.BadZipFile as e:
        raise RuntimeError(
            f"Content at {url} isn't a valid .xlsx workbook — most likely an "
            "HTML sign-in or 'request access' page, i.e. the file isn't shared "
            "as 'Anyone with the link: Viewer'. Check sharing, or pass a local "
            "file instead."
        ) from e


def sheet_row_dicts(ws) -> list[dict[str, Any]]:
    """One dict per non-blank data row, keyed by lowercase header name."""
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return []
    header_index = {
        str(h).strip().lower(): i for i, h in enumerate(header) if h is not None
    }
    rows = []
    for row in rows_iter:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(
            {
                name: (row[i] if i < len(row) else None)
                for name, i in header_index.items()
            }
        )
    return rows


def is_blank(value: Any) -> bool:
    """None, empty, or a formula error — all mean "no value here"."""
    if value is None:
        return True
    if isinstance(value, str):
        stripped = value.strip()
        return stripped == "" or stripped in EXCEL_ERRORS
    return False


def to_text(value: Any) -> Optional[str]:
    """TEXT columns. openpyxl returns numeric-looking cells as float, and
    binding a float into TEXT has no implicit Postgres cast — so whole
    numbers are stringified without the trailing '.0' (e.g. an id column
    typed as a number in the sheet)."""
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def to_int(value: Any) -> Optional[int]:
    if is_blank(value):
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def to_float(value: Any) -> Optional[float]:
    if is_blank(value):
        return None
    try:
        return float(str(value).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def to_bool(value: Any) -> Optional[bool]:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().upper()
    if text in _BOOL_TRUE:
        return True
    if text in _BOOL_FALSE:
        return False
    try:  # numeric-ish leftovers, e.g. "2.0"
        return bool(float(text))
    except ValueError:
        return None


def to_clock(value: Any) -> Optional[str]:
    """Clock time → the schema's HH:MM TEXT convention."""
    if is_blank(value):
        return None
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    return str(value).strip() or None


def to_duration(value: Any) -> Optional[str]:
    """Duration → HH:MM with hours NOT wrapped at 24 (a multi-day trip
    duration like 25:03 is valid, which datetime.time cannot represent)."""
    if is_blank(value):
        return None
    if isinstance(value, datetime.timedelta):
        hours, minutes = divmod(round(value.total_seconds() / 60), 60)
        return f"{hours:02d}:{minutes:02d}"
    return str(value).strip() or None


def to_date(value: Any) -> Optional[datetime.date]:
    if is_blank(value):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None
